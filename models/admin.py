# models/admin.py

from django.contrib import admin
from django.urls import path
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
import openpyxl
from io import BytesIO
from .models import UserModel, ComeGoneTimeModel


@admin.register(UserModel)
class UserModelAdmin(admin.ModelAdmin):
    list_display = ('user', 'full_name', 'status')
    list_filter = ('status',)
    search_fields = ('user', 'full_name')
    ordering = ('user',)


@admin.register(ComeGoneTimeModel)
class ComeGoneTimeAdmin(admin.ModelAdmin):
    list_display = ('get_user_id_display', 'get_user_full_name_display', 'get_formatted_time_display_admin',
                    # Admin panel uchun alohida metod
                    'get_event_status_display')
    list_filter = ('status', 'user', ('time', admin.DateFieldListFilter))
    search_fields = ('user__user', 'user__full_name')
    ordering = ['-time']

    change_list_template = "admin/models/comegonetimemodel/change_list.html"

    @admin.display(description='User ID', ordering='user__user')
    def get_user_id_display(self, obj):
        return obj.user.user

    @admin.display(description='To\'liq Ism', ordering='user__full_name')
    def get_user_full_name_display(self, obj):
        return obj.user.full_name or "Noma'lum"

    @admin.display(description='Sana va Vaqt (Lokal)', ordering='time')
    def get_formatted_time_display_admin(self, obj):
        local_time = timezone.localtime(obj.time)
        return local_time.strftime('%d-%m-%Y %H:%M:%S')

    @admin.display(description='Hodisa Holati', ordering='status')
    def get_event_status_display(self, obj):
        if obj.status is True:
            return "Keldi"
        elif obj.status is False:
            return "Ketdi"
        return "Noma'lum"

    def get_urls(self):
        urls = super().get_urls()
        info = self.model._meta
        url_name = f'{info.app_label}_{info.model_name}_export_all_excel'
        custom_urls = [
            path(
                'export-excel/',
                self.admin_site.admin_view(self.export_all_to_excel),
                name=url_name
            ),
        ]
        return custom_urls + urls

    def export_all_to_excel(self, request):
        queryset = ComeGoneTimeModel.objects.select_related('user').order_by('-time').all()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = str(_('Keldi-Ketdi Hisoboti'))

        headers = [
            str(_('User ID')),
            str(_('To\'liq Ism')),
            str(_('Sana')),
            str(_('Vaqt')),
            str(_('Hodisa Holati'))
        ]

        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = openpyxl.styles.Font(bold=True)

        for row_num, item in enumerate(queryset.iterator(), 2):
            local_event_time = timezone.localtime(item.time)
            event_date = local_event_time.strftime('%d-%m-%Y')
            event_time_str = local_event_time.strftime('%H:%M:%S')

            status_display = str(_("Noma'lum"))
            if item.status is True:
                status_display = str(_("Keldi"))
            elif item.status is False:
                status_display = str(_("Ketdi"))

            row_data = [
                item.user.user,
                item.user.full_name or str(_("Noma'lum")),
                event_date,
                event_time_str,
                status_display,
            ]
            for col_num, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_str_len = len(str(cell.value))
                        if cell_str_len > max_length:
                            max_length = cell_str_len
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > 0 else 12
            worksheet.column_dimensions[column_letter].width = adjusted_width

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        response = HttpResponse(
            excel_buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        current_time_str = timezone.localtime(timezone.now()).strftime(
            '%Y-%m-%d_%H%M%S')
        filename = f"keldi_ketdi_hisoboti_barchasi_{current_time_str}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response